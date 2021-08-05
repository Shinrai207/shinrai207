import openpyxl
wb = openpyxl.load_workbook("0805.xlsx")
ws_hello = wb["hello"]
ws_hello2 = wb["hello2"]
  
for i in range (1, 9):
    for j in range (1, 3):
      
        c = ws_hello.cell(row = i, column = j)
  
        ws_hello2.cell(row = i, column = j).value = c.value
  
wb.save("0805.xlsx")